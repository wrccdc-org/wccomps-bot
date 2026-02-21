"""Import qualifier scores from Excel spreadsheet."""

from decimal import Decimal
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError, CommandParser
from django.db import transaction
from registration.models import Event, EventTeamAssignment, Season, TeamRegistration

from scoring.models import (
    IncidentReport,
    InjectGrade,
    OrangeTeamBonus,
    RedTeamFinding,
    ServiceDetail,
    ServiceScore,
)
from team.models import Team


def _parse_team_num(label: object) -> int | None:
    """Parse team number from a label like 'Team 01' or 'Team 1'."""
    if not label or not str(label).startswith("Team"):
        return None
    return int(str(label).replace("Team ", "").strip())


def _to_decimal(value: object) -> Decimal:
    """Convert a cell value to Decimal, defaulting to 0."""
    if value is None:
        return Decimal("0")
    return Decimal(str(value))


class Command(BaseCommand):
    """Import qualifier scores from an Excel spreadsheet."""

    help = "Import qualifier scores from an Excel spreadsheet"

    def add_arguments(self, parser: CommandParser) -> None:
        """Add command arguments."""
        parser.add_argument("xlsx_path", type=str, help="Path to the Excel file")
        parser.add_argument("--event-name", default="2026 Qualifier", help="Event name")
        parser.add_argument("--season-year", type=int, default=2026, help="Season year")
        parser.add_argument("--dry-run", action="store_true", help="Preview without saving")

    @transaction.atomic
    def handle(self, *args: str, **options: object) -> None:
        """Execute the import."""
        xlsx_path = Path(str(options["xlsx_path"]))
        if not xlsx_path.exists():
            raise CommandError(f"File not found: {xlsx_path}")

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("DRY RUN - no changes will be saved"))

        # 1. Create Season and Event
        season, _ = Season.objects.get_or_create(
            year=options["season_year"],
            defaults={"name": f"{options['season_year']} Season", "is_active": True},
        )
        event, _ = Event.objects.get_or_create(
            season=season,
            event_type="qualifier",
            defaults={
                "name": options["event_name"],
                "date": f"{options['season_year']}-02-08",
                "is_finalized": True,
            },
        )
        self.stdout.write(f"Event: {event}")

        # 2. Create teams from Team Number Assignments
        ws_teams = wb["Team Number Assignments"]
        teams: dict[int, Team] = {}
        for row in range(2, ws_teams.max_row + 1):
            team_label = ws_teams.cell(row=row, column=1).value
            team_name = ws_teams.cell(row=row, column=3).value
            if not team_label or not team_name:
                continue
            team_num = _parse_team_num(team_label)
            if team_num is None:
                continue
            team_obj, created = Team.objects.get_or_create(
                team_number=team_num,
                defaults={"team_name": str(team_name)},
            )
            if not created:
                team_obj.team_name = str(team_name)
                team_obj.save(update_fields=["team_name"])
            teams[team_num] = team_obj
            if created:
                self.stdout.write(f"  Created Team {team_num}: {team_name}")

        # Also create teams from Rankings that may not be in assignments (28, 29)
        ws_rankings = wb["Rankings & Totals"]
        for row in range(2, ws_rankings.max_row + 1):
            team_num_val = ws_rankings.cell(row=row, column=1).value
            team_name_val = ws_rankings.cell(row=row, column=2).value
            if team_num_val is None or team_name_val is None:
                continue
            team_num = int(float(str(team_num_val)))
            if team_num not in teams:
                team_obj, created = Team.objects.get_or_create(
                    team_number=team_num,
                    defaults={"team_name": str(team_name_val)},
                )
                teams[team_num] = team_obj
                if created:
                    self.stdout.write(f"  Created Team {team_num}: {team_name_val}")

        self.stdout.write(f"  {len(teams)} teams loaded")

        # 3. Create registrations and event team assignments
        for team_obj in teams.values():
            reg, _ = TeamRegistration.objects.get_or_create(
                school_name=team_obj.team_name,
                defaults={"status": "credentials_sent"},
            )
            EventTeamAssignment.objects.get_or_create(
                event=event,
                team=team_obj,
                defaults={"registration": reg},
            )

        # 4. Import Service Scores
        ws_svc = wb["Total Service Points"]
        svc_imported = 0
        for row in range(2, ws_svc.max_row + 1):
            team_num = _parse_team_num(ws_svc.cell(row=row, column=1).value)
            if team_num is None or team_num not in teams:
                continue
            svc_pts = _to_decimal(ws_svc.cell(row=row, column=2).value)
            sla = _to_decimal(ws_svc.cell(row=row, column=3).value)
            pt_adj = _to_decimal(ws_svc.cell(row=row, column=4).value)
            ServiceScore.objects.update_or_create(
                team=teams[team_num],
                defaults={
                    "event": event,
                    "service_points": svc_pts,
                    "sla_violations": sla,
                    "point_adjustments": pt_adj,
                },
            )
            svc_imported += 1
        self.stdout.write(f"  Imported {svc_imported} service scores")

        # 5. Import per-service details
        ws_detail = wb["Service Points Per Service"]
        svc_detail_count = 0
        for svc_row in range(2, ws_detail.max_row + 1):
            svc_name = ws_detail.cell(row=svc_row, column=2).value
            if not svc_name or svc_name == "Total":
                continue
            for team_num in sorted(teams.keys()):
                col = team_num + 2  # Team 01 = col C (3)
                pts = ws_detail.cell(row=svc_row, column=col).value
                if pts is None:
                    pts = 0
                ServiceDetail.objects.update_or_create(
                    team=teams[team_num],
                    event=event,
                    service_name=str(svc_name),
                    defaults={"points": Decimal(str(pts))},
                )
                svc_detail_count += 1
        self.stdout.write(f"  Imported {svc_detail_count} service detail records")

        # 6. Import Inject totals
        ws_inj = wb["Inject Points"]
        inj_count = 0
        for row in range(2, ws_inj.max_row + 1):
            team_num = _parse_team_num(ws_inj.cell(row=row, column=1).value)
            if team_num is None or team_num not in teams:
                continue
            raw_total = ws_inj.cell(row=row, column=20).value  # col T
            if raw_total is None or raw_total == 0:
                continue
            InjectGrade.objects.update_or_create(
                team=teams[team_num],
                inject_id="qualifier-total",
                defaults={
                    "event": event,
                    "inject_name": "Qualifier Total",
                    "points_awarded": Decimal(str(raw_total)),
                    "is_approved": True,
                },
            )
            inj_count += 1
        self.stdout.write(f"  Imported {inj_count} inject grades")

        # 7. Import Orange scores
        ws_orange = wb["Orange Team Scores"]
        orange_count = 0
        for row in range(2, ws_orange.max_row + 1):
            team_num = _parse_team_num(ws_orange.cell(row=row, column=1).value)
            if team_num is None or team_num not in teams:
                continue
            raw_pts = ws_orange.cell(row=row, column=2).value
            if raw_pts is None or raw_pts == 0:
                continue
            OrangeTeamBonus.objects.update_or_create(
                team=teams[team_num],
                description="Qualifier orange team score",
                defaults={
                    "event": event,
                    "points_awarded": Decimal(str(raw_pts)),
                    "is_approved": True,
                },
            )
            orange_count += 1
        self.stdout.write(f"  Imported {orange_count} orange scores")

        # 8. Import Red Team Deductions
        # Create one finding per team per category because deduction
        # amounts can vary by team (e.g. Realm Persistence).
        ws_red = wb["Red Team Deductions"]
        categories: dict[int, str] = {}
        for col in range(3, ws_red.max_column + 1):
            header = ws_red.cell(row=1, column=col).value
            if header is None or header == 0:
                continue
            header_str = str(header)
            if header_str in ("Total:", "Points Back"):
                continue
            categories[col] = header_str

        red_count = 0
        for row in range(2, ws_red.max_row + 1):
            team_num = _parse_team_num(ws_red.cell(row=row, column=1).value)
            if team_num is None or team_num not in teams:
                continue

            for col, cat_name in categories.items():
                pts = ws_red.cell(row=row, column=col).value
                if pts is None or pts == 0 or pts == 0.0:
                    continue
                finding, created = RedTeamFinding.objects.update_or_create(
                    event=event,
                    notes=f"Qualifier import: {cat_name} (Team {team_num})",
                    defaults={
                        "attack_vector": cat_name,
                        "points_per_team": abs(Decimal(str(pts))),
                        "is_approved": True,
                    },
                )
                finding.affected_teams.add(teams[team_num])
                red_count += 1

        self.stdout.write(f"  Imported {red_count} red team deduction entries")

        # 8b. Import "Points Back" as IncidentReport (incident recovery points)
        points_back_col: int | None = None
        for col in range(3, ws_red.max_column + 1):
            if ws_red.cell(row=1, column=col).value == "Points Back":
                points_back_col = col
                break

        pb_count = 0
        if points_back_col:
            for row in range(2, ws_red.max_row + 1):
                team_num = _parse_team_num(ws_red.cell(row=row, column=1).value)
                if team_num is None or team_num not in teams:
                    continue
                pb_pts = ws_red.cell(row=row, column=points_back_col).value
                if pb_pts is None or pb_pts == 0 or pb_pts == 0.0:
                    continue
                IncidentReport.objects.update_or_create(
                    team=teams[team_num],
                    attack_description="Qualifier import: Points Back",
                    defaults={
                        "event": event,
                        "source_ip": "127.0.0.1",
                        "attack_detected_at": f"{options['season_year']}-02-08T00:00:00Z",
                        "attack_mitigated": True,
                        "gold_team_reviewed": True,
                        "points_returned": Decimal(str(pb_pts)),
                        "reviewer_notes": "Imported from qualifier spreadsheet",
                    },
                )
                pb_count += 1
        self.stdout.write(f"  Imported {pb_count} incident recovery (points back) records")

        # 9. Recalculate final scores from source records
        from scoring.calculator import recalculate_all_scores

        recalculate_all_scores()
        self.stdout.write(self.style.SUCCESS("Recalculated final scores from imported data"))

        if options["dry_run"]:
            raise CommandError("DRY RUN - rolling back all changes")
