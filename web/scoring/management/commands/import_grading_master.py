"""Import grading master inject scores and judge comments from exported XLSX."""

from __future__ import annotations

import re
from decimal import Decimal
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError, CommandParser
from django.db import transaction
from openpyxl.worksheet.worksheet import Worksheet
from registration.models import Event, Season

from scoring.models import InjectScore
from team.models import Team

# Inject columns on the summary sheet (0-indexed from the first inject column).
# Maps inject number to (short_name, max_points).
INJECT_MAP: dict[int, tuple[str, int]] = {
    0: ("Inject 00 - Welcome", 100),
    1: ("Inject 01 - Frappe ERPNext 01 - Vulnerability Hunt", 200),
    2: ("Inject 02 - Onboarding", 150),
    3: ("Inject 03 - Domain Controller", 300),
    4: ("Inject 04 - IR Policy", 150),
    5: ("Inject 05 - Frappe ERPNext 02 - Identify Exposed Files", 200),
    6: ("Inject 06 - Wazuh 01 - Create Dashboard", 200),
    7: ("Inject 07 - Video", 300),
    8: ("Inject 08 - Offboarding", 200),
    9: ("Inject 09 - Frappe ERPNext 03 - Identify Real Time Attack Pattern", 250),
    10: ("Inject 10 - Create Server Baselines", 200),
    11: ("Inject 11 - Wazuh 02 - Expand Dashboard", 200),
    12: ("Inject 12 - AD Server Vulnerability Scan", 170),
    13: ("Inject 13 - Mascot", 140),
    14: ("Inject 14 - Wazuh 03 - Dashboard Monitor", 200),
    15: ("Inject 15 - Feedback", 100),
}

_TEAM_RE = re.compile(r"^Team\s+(\d+)$")


def _parse_team_num(label: object) -> int | None:
    """Parse team number from a label like 'Team 01' or 'Team 1'."""
    if not label:
        return None
    m = _TEAM_RE.match(str(label).strip())
    return int(m.group(1)) if m else None


def _to_decimal(value: object) -> Decimal:
    """Convert a cell value to Decimal, defaulting to 0."""
    if value is None:
        return Decimal("0")
    return Decimal(str(value))


def _find_inject_number(sheet_name: str) -> int | None:
    """Extract inject number from a sheet tab name.

    Looks for patterns like 'Inject 01', 'Inject 1', '01 -', etc.
    """
    m = re.search(r"(?:Inject\s*)?(\d{1,2})", sheet_name)
    if m:
        num = int(m.group(1))
        if 0 <= num <= 15:
            return num
    return None


def _parse_summary_sheet(
    ws: Worksheet,
    teams: dict[int, Team],
) -> dict[int, dict[int, Decimal]]:
    """Parse the summary sheet to get averaged inject totals per team.

    Returns: {team_num: {inject_num: points_awarded}}

    The summary sheet has:
    - Row 1: headers with inject names
    - Row 2: "Max Score" row with max points per inject
    - Rows 3+: Team rows with "Team N" in col A, inject scores in subsequent cols
    - The first inject column is column B (index 2), last inject col is Q (index 17)
    - Column R (index 18) has "Totals"
    """
    scores: dict[int, dict[int, Decimal]] = {}

    # Find the inject columns by scanning the header/max-score area.
    # The summary has 16 inject columns starting from column B.
    inject_start_col = 2  # Column B

    for row in range(1, ws.max_row + 1):
        team_num = _parse_team_num(ws.cell(row=row, column=1).value)
        if team_num is None or team_num not in teams:
            continue

        team_scores: dict[int, Decimal] = {}
        for inject_num in range(16):
            col = inject_start_col + inject_num
            val = _to_decimal(ws.cell(row=row, column=col).value)
            team_scores[inject_num] = val
        scores[team_num] = team_scores

    return scores


def _find_judge_sections(ws: Worksheet) -> list[tuple[str, int, int]]:
    """Find judge sections in a detail sheet.

    Each detail sheet contains multiple judge sections stacked vertically.
    A judge section starts with the judge name alone in column A on the
    same row as the rubric headers (including a "Comments" column), followed
    by a "Max Score" row, then team data rows.

    Returns: [(judge_name, header_row, comments_col), ...]
    """
    sections: list[tuple[str, int, int]] = []
    for row in range(1, ws.max_row + 1):
        a_val = ws.cell(row=row, column=1).value
        if a_val is None:
            continue
        s = str(a_val).strip()
        if s.startswith("Team") or s == "Max Score":
            continue
        # Potential judge name — check if this row has a "Comments" header
        comments_col: int | None = None
        for col in range(2, ws.max_column + 1):
            h = ws.cell(row=row, column=col).value
            if h and str(h).strip().lower() == "comments":
                comments_col = col
                break
        if comments_col is not None:
            sections.append((s, row, comments_col))
    return sections


def _parse_detail_sheets(
    wb: openpyxl.Workbook,
    teams: dict[int, Team],
) -> dict[int, dict[int, list[str]]]:
    """Parse per-inject detail sheets for judge comments.

    Returns: {inject_num: {team_num: [comment1, comment2, ...]}}

    Each detail sheet has multiple judge sections stacked vertically.
    The judge name appears in column A on the header row, with rubric
    criteria and a "Comments" column. Team rows follow with comments.
    """
    comments: dict[int, dict[int, list[str]]] = {}

    # Skip the first sheet (summary)
    for ws in wb.worksheets[1:]:
        inject_num = _find_inject_number(ws.title)
        if inject_num is None:
            continue

        if inject_num not in comments:
            comments[inject_num] = {}

        sections = _find_judge_sections(ws)
        for _judge_name, header_row, comments_col in sections:
            # Team rows start 2 rows after the header (header, Max Score, teams)
            for row in range(header_row + 2, ws.max_row + 1):
                team_num = _parse_team_num(ws.cell(row=row, column=1).value)
                if team_num is None:
                    # End of this judge section (blank row or next section)
                    if ws.cell(row=row, column=1).value is None:
                        break
                    continue
                if team_num not in teams:
                    continue

                comment = ws.cell(row=row, column=comments_col).value
                if comment and str(comment).strip():
                    comment_text = str(comment).strip()
                    if comment_text.lower() == "no submission":
                        continue
                    if team_num not in comments[inject_num]:
                        comments[inject_num][team_num] = []
                    comments[inject_num][team_num].append(comment_text)

    return comments


class Command(BaseCommand):
    """Import grading master inject scores and judge comments from XLSX."""

    help = "Import grading master inject scores from exported XLSX"

    def add_arguments(self, parser: CommandParser) -> None:
        """Add command arguments."""
        parser.add_argument("xlsx_path", type=str, help="Path to the Grading Master XLSX")
        parser.add_argument("--event-name", default="2026 Qualifier", help="Event name")
        parser.add_argument("--season-year", type=int, default=2026, help="Season year")
        parser.add_argument("--dry-run", action="store_true", help="Preview without saving")

    @transaction.atomic
    def handle(self, *args: str, **options: object) -> None:
        """Execute the import."""
        xlsx_path = Path(str(options["xlsx_path"]))
        if not xlsx_path.exists():
            raise CommandError(f"File not found: {xlsx_path}")

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("DRY RUN - no changes will be saved"))

        # 1. Get event
        season_year = int(str(options["season_year"]))
        season = Season.objects.get(year=season_year)
        event = Event.objects.get(season=season, event_type="qualifier")
        self.stdout.write(f"Event: {event}")

        # 2. Build team lookup from DB (only active teams with scores > 0)
        teams: dict[int, Team] = {t.team_number: t for t in Team.objects.all()}
        self.stdout.write(f"  {len(teams)} teams in database")

        # 3. Parse summary sheet for averaged inject totals
        ws_summary = wb.worksheets[0]
        inject_scores = _parse_summary_sheet(ws_summary, teams)
        self.stdout.write(f"  Parsed scores for {len(inject_scores)} teams from summary")

        # 4. Parse detail sheets for judge comments
        all_comments = _parse_detail_sheets(wb, teams)
        comment_inject_count = sum(1 for inj in all_comments.values() if inj)
        self.stdout.write(f"  Found comments for {comment_inject_count} injects")

        # 5. Create/update InjectScore records
        created_count = 0
        updated_count = 0
        for team_num, team_injects in inject_scores.items():
            team = teams[team_num]
            for inject_num, points in team_injects.items():
                inject_id = f"qual-{inject_num:02d}"
                inject_name, max_points = INJECT_MAP[inject_num]

                # Merge comments from all judges
                raw_comments = ""
                if inject_num in all_comments and team_num in all_comments[inject_num]:
                    raw_comments = " | ".join(all_comments[inject_num][team_num])

                _, created = InjectScore.objects.update_or_create(
                    team=team,
                    inject_id=inject_id,
                    defaults={
                        "event": event,
                        "inject_name": inject_name,
                        "points_awarded": points,
                        "max_points": Decimal(str(max_points)),
                        "notes": raw_comments,
                        "feedback": "",
                        "feedback_approved": False,
                        "is_approved": True,
                    },
                )
                if created:
                    created_count += 1
                else:
                    updated_count += 1

        self.stdout.write(f"  Created {created_count}, updated {updated_count} inject scores")

        # 6. Delete old qualifier-total records
        deleted, _ = InjectScore.objects.filter(
            inject_id="qualifier-total",
            event=event,
        ).delete()
        if deleted:
            self.stdout.write(f"  Deleted {deleted} old qualifier-total records")

        # 7. Recalculate final scores
        from scoring.calculator import recalculate_all_scores

        recalculate_all_scores()
        self.stdout.write(self.style.SUCCESS("Recalculated final scores from imported data"))

        if options["dry_run"]:
            raise CommandError("DRY RUN - rolling back all changes")
